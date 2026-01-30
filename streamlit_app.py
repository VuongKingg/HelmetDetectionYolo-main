import streamlit as st
import cv2
import tempfile
import time
import pandas as pd
import os
import numpy as np
from datetime import datetime
from app.detector import HelmetDetector
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# --- CẤU HÌNH HỆ THỐNG ---
CONFIDENCE_THRESHOLD = 0.5
EXCEL_FILE = 'report_vi_pham.xlsx'
EVIDENCE_DIR = 'evidence_images'

if not os.path.exists(EVIDENCE_DIR):
    os.makedirs(EVIDENCE_DIR)

# --- CLASS THEO DÕI ĐỐI TƯỢNG (TRACKER) ---
class Tracker:
    def __init__(self, max_disappeared=40, max_distance=100):
        self.nextObjectID = 0
        self.objects = {}
        self.disappeared = {}
        self.max_disappeared = max_disappeared
        self.max_distance = max_distance

    def register(self, centroid):
        self.objects[self.nextObjectID] = centroid
        self.disappeared[self.nextObjectID] = 0
        self.nextObjectID += 1
        return self.nextObjectID - 1

    def deregister(self, objectID):
        del self.objects[objectID]
        del self.disappeared[objectID]

    def update(self, rects):
        if len(rects) == 0:
            for objectID in list(self.disappeared.keys()):
                self.disappeared[objectID] += 1
                if self.disappeared[objectID] > self.max_disappeared:
                    self.deregister(objectID)
            return self.objects

        inputCentroids = np.zeros((len(rects), 2), dtype="int")
        for (i, (startX, startY, endX, endY)) in enumerate(rects):
            cX = int((startX + endX) / 2.0)
            cY = int((startY + endY) / 2.0)
            inputCentroids[i] = (cX, cY)

        if len(self.objects) == 0:
            for i in range(0, len(inputCentroids)):
                self.register(inputCentroids[i])
        else:
            objectIDs = list(self.objects.keys())
            objectCentroids = list(self.objects.values())

            D = []
            for oc in objectCentroids:
                row = []
                for ic in inputCentroids:
                    dist = np.linalg.norm(np.array(oc) - np.array(ic))
                    row.append(dist)
                D.append(row)
            D = np.array(D)

            rows = D.min(axis=1).argsort()
            cols = D.argmin(axis=1)[rows]
            usedRows = set()
            usedCols = set()

            for (row, col) in zip(rows, cols):
                if row in usedRows or col in usedCols: continue
                if D[row, col] > self.max_distance: continue
                objectID = objectIDs[row]
                self.objects[objectID] = inputCentroids[col]
                self.disappeared[objectID] = 0
                usedRows.add(row)
                usedCols.add(col)

            unusedRows = set(range(0, D.shape[0])).difference(usedRows)
            unusedCols = set(range(0, D.shape[1])).difference(usedCols)

            for row in unusedRows:
                objectID = objectIDs[row]
                self.disappeared[objectID] += 1
                if self.disappeared[objectID] > self.max_disappeared:
                    self.deregister(objectID)
            for col in unusedCols:
                self.register(inputCentroids[col])

        return self.objects

# --- HÀM TRANG TRÍ EXCEL ---
def format_excel_file(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        wb.save(filename)
    except Exception as e:
        print(f"Lỗi format excel: {e}")

# --- HÀM GHI LOG ---
def log_violation(object_id, image_path):
    now = datetime.now()
    # LƯU Ý: Tên cột ở đây phải khớp với tên cột lúc gọi hiển thị
    new_data = {
        "ID Vi Phạm": [object_id],
        "Thời gian": [now.strftime("%Y-%m-%d %H:%M:%S")],
        "Loại lỗi": ["Không đội mũ bảo hiểm"],
        "Đường dẫn ảnh": [image_path]
    }
    df_new = pd.DataFrame(new_data)
    
    if os.path.exists(EXCEL_FILE):
        try:
            df_old = pd.read_excel(EXCEL_FILE)
            if object_id not in df_old['ID Vi Phạm'].values:
                df_combined = pd.concat([df_old, df_new], ignore_index=True)
                df_combined.to_excel(EXCEL_FILE, index=False)
                format_excel_file(EXCEL_FILE)
                return True
        except:
            # Nếu file lỗi format cũ, tạo lại mới
            df_new.to_excel(EXCEL_FILE, index=False)
            format_excel_file(EXCEL_FILE)
            return True
    else:
        df_new.to_excel(EXCEL_FILE, index=False)
        format_excel_file(EXCEL_FILE)
        return True
    return False

# --- GIAO DIỆN CHÍNH ---
st.set_page_config(page_title="AI Traffic Monitor", layout="wide", page_icon="📹")
st.title("🪖 Hệ thống Giám sát Vi phạm Mũ bảo hiểm")
st.markdown("---")

@st.cache_resource
def init_detector():
    return HelmetDetector()

try:
    detector = init_detector()
    tracker = Tracker(max_disappeared=40, max_distance=100)
except Exception as e:
    st.error(f"Lỗi khởi tạo: {e}")
    st.stop()

col1, col2 = st.columns([0.6, 0.4])

with col1:
    st.subheader("📡 Camera Giám sát")
    source_type = st.radio("Nguồn dữ liệu:", ["Video Upload", "Webcam"], horizontal=True)
    
    cap = None
    if source_type == "Video Upload":
        video_file = st.file_uploader("Chọn video MP4/AVI", type=['mp4', 'avi', 'mov'])
        if video_file:
            tfile = tempfile.NamedTemporaryFile(delete=False)
            tfile.write(video_file.read())
            cap = cv2.VideoCapture(tfile.name)
    else:
        cap = cv2.VideoCapture(0)

    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        start_btn = st.button("▶️ Bắt đầu chạy", type="primary", use_container_width=True)
    with col_btn2:
        stop_btn = st.button("⏹️ Dừng hệ thống", use_container_width=True)
    
    st_frame = st.empty()

with col2:
    st.subheader("📋 Nhật ký & Bằng chứng")
    placeholder_table = st.empty()
    
logged_ids = set()

# --- LOOP XỬ LÝ ---
if start_btn and cap:
    # Tạo các placeholder để giữ chỗ hiển thị
    status_text_placeholder = col2.empty() # Nơi hiển thị trạng thái "Đang ghi..."
    live_table_placeholder = col2.empty()  # Nơi hiển thị bảng danh sách
    
    while cap.isOpened() and not stop_btn:
        ret, frame = cap.read()
        if not ret: break

        # 1. Detection (Giữ nguyên logic của bạn)
        results = detector.model.predict(frame, conf=CONFIDENCE_THRESHOLD, verbose=False)
        result = results[0]

        violation_rects = []
        for box in result.boxes:
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            cls = int(box.cls[0])
            label = detector.model.names[cls]
            
            # QUAN TRỌNG: Kiểm tra đúng nhãn để bắt lỗi
            # Nếu model của bạn trả về 'No Helmet' hoặc 'without helmet' (tùy model)
            if label in ["without helmet", "No Helmet", "no helmet"]: 
                violation_rects.append((x1, y1, x2, y2))
                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 2)
                cv2.putText(frame, "KHONG MU", (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
            else:
                # Vẽ màu xanh cho các trường hợp khác
                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)

        # 2. Tracking Update
        objects = tracker.update(violation_rects)
        
        # Biến cờ để kiểm tra có lỗi mới trong frame này không
        has_new_violation = False 

        for (objectID, centroid) in objects.items():
            # Vẽ ID lên video
            text = f"ID {objectID}"
            cv2.putText(frame, text, (centroid[0], centroid[1]-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
            cv2.circle(frame, (centroid[0], centroid[1]), 4, (0, 0, 255), -1)

            # LOGIC GHI DỮ LIỆU
            if objectID not in logged_ids:
                img_name = f"violation_{objectID}_{int(time.time())}.jpg"
                save_path = os.path.join(EVIDENCE_DIR, img_name)
                cv2.imwrite(save_path, frame)
                
                # Gọi hàm log và kiểm tra kết quả trả về
                if log_violation(objectID, save_path):
                    logged_ids.add(objectID)
                    has_new_violation = True # Đánh dấu có dữ liệu mới
                    st.toast(f"🚨 Đã ghi nhận vi phạm: ID {objectID}", icon="⚠️")

        # 3. Hiển thị Video
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        st_frame.image(frame_rgb, channels="RGB")

        # 4. CẬP NHẬT BẢNG REAL-TIME (CẢI TIẾN)
        # Chỉ đọc lại file Excel nếu có vi phạm mới hoặc mỗi 30 frame (để giảm lag)
        # Tuy nhiên, để chắc chắn hiển thị, ta sẽ force update nếu 'has_new_violation' = True
        
        if os.path.exists(EXCEL_FILE):
            try:
                # Đọc dữ liệu mới nhất
                df_display = pd.read_excel(EXCEL_FILE)
                
                # Sắp xếp mới nhất lên đầu
                df_display = df_display.sort_values(by="Thời gian", ascending=False)
                
                # Lấy 3 cột quan trọng nhất & 5 dòng đầu tiên
                df_mini = df_display[['ID Vi Phạm', 'Thời gian', 'Loại lỗi']].head(5)
                
                # Cập nhật trạng thái
                status_text_placeholder.markdown(f"**⚡ Trạng thái:** Đã ghi nhận **{len(df_display)}** vi phạm.")
                
                # Hiển thị bảng (Dùng container để vẽ đè lên cũ)
                with live_table_placeholder.container():
                     # Sử dụng HTML Table đơn giản để tránh lỗi Duplicate Key của Streamlit dataframe
                     st.table(df_mini)

            except Exception as e:
                print(f"Lỗi đọc bảng: {e}")

    cap.release()

# --- PHẦN XEM LẠI LỊCH SỬ (KHI DỪNG VIDEO) ---
if os.path.exists(EXCEL_FILE):
    st.divider()
    st.subheader("🔍 Tra cứu lịch sử vi phạm")
    
    try:
        df_review = pd.read_excel(EXCEL_FILE)
        df_review = df_review.sort_values(by="Thời gian", ascending=False)

        c1, c2 = st.columns([0.6, 0.4])
        with c1:
            st.info("Click vào hàng để xem ảnh:")
            event = st.dataframe(
                df_review,
                column_config={
                    "Đường dẫn ảnh": st.column_config.TextColumn("Link ảnh"),
                },
                use_container_width=True,
                hide_index=True,
                on_select="rerun",
                selection_mode="single-row",
                key="history_static"
            )
        
        with c2:
            if len(event.selection.rows) > 0:
                idx = event.selection.rows[0]
                img_path = df_review.iloc[idx]["Đường dẫn ảnh"]
                v_id = df_review.iloc[idx]["ID Vi Phạm"]
                st.warning(f"📸 Bằng chứng ID: {v_id}")
                if os.path.exists(img_path):
                    st.image(img_path, use_container_width=True)
                else:
                    st.error("Ảnh không tồn tại.")
    except Exception as e:
        st.error("File dữ liệu đang bị lỗi hoặc trống.")